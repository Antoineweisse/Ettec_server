from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment,Border,Side
from openpyxl.utils import get_column_letter
from io import BytesIO

VIOLET = "b44de0"
ORANGE = "ffa500"
BLEU_CLAIR = "77b5fe"
GREEN = "008020"
BLEU = "0080ff"
YELLOW = "ffff00"
RED ="ff0000"

def makeDate(ws):
    fill = PatternFill(start_color=VIOLET, end_color=VIOLET, fill_type="solid")
    ws.column_dimensions["A"].width = 16

    ws["A1"] = "Jours"
    ws["A2"] = "Dates"
    ws["A1"].font = Font(bold=True,color="000000")
    ws["A2"].font = Font(bold=True,color="000000")
    ws["A1"].fill = fill
    ws["A2"].fill = fill

    for i in range(2, 33):
        col_letter = get_column_letter(i)
        ws[f"{col_letter}1"] = str(i-1)
        ws.merge_cells(f"{col_letter}1:{col_letter}2")
        ws[f"{col_letter}1"].alignment = Alignment(horizontal="center", vertical="bottom")
        ws.column_dimensions[col_letter].width = 3
        ws[f"{col_letter}1"].font = Font(bold=True,color="000000")
        ws[f"{col_letter}1"].fill = fill

    fill = PatternFill(start_color=ORANGE, end_color=ORANGE, fill_type="solid")
    col_letter = get_column_letter(33)
    ws[f"{col_letter}1"] = "Totaux"
    ws.merge_cells(f"{col_letter}1:{col_letter}2")
    ws[f"{col_letter}1"].alignment = Alignment(horizontal="left", vertical="bottom")
    ws.column_dimensions[col_letter].width = 7
    ws[f"{col_letter}1"].font = Font(bold=True)
    ws[f"{col_letter}1"].fill = fill

    fill = PatternFill(start_color=YELLOW, end_color=YELLOW, fill_type="solid")
    col_letter = get_column_letter(34)
    ws[f"{col_letter}1"] = "Accomptes - Primes - Divers"
    ws.column_dimensions[col_letter].width = 20
    ws.column_dimensions[col_letter].height = 100
    ws.merge_cells(f"{col_letter}1:{col_letter}2")
    ws[f"{col_letter}1"].alignment = Alignment(horizontal="left", vertical="bottom", wrap_text=True)
    ws[f"{col_letter}1"].font = Font(bold=False)
    ws[f"{col_letter}1"].fill = fill

def makeTitle(ws,title):
    ws.merge_cells("A3:AH4")
    ws["A3"] = title
    ws["A3"].alignment = Alignment(horizontal="center", vertical="center")
    ws["A3"].font = Font(bold=True, size=14)
    fill = PatternFill(start_color=BLEU_CLAIR, end_color=BLEU_CLAIR, fill_type="solid")
    ws["A3"].fill = fill

def makeEmployeeName(ws,Name, contrat,debut):
    fill = PatternFill(start_color=GREEN, end_color=GREEN, fill_type="solid")
    cell1 = f"A{debut}"
    cell2 = f"A{debut+1}"

    ws.merge_cells(f"{cell1}:AH{debut}")
    ws.merge_cells(f"{cell2}:AH{debut+1}")

    ws[cell1] = Name
    ws[cell1].alignment = Alignment(horizontal="center", vertical="center")
    ws[cell1].font = Font(bold=True, size=12,color="000000")
    ws[cell1].fill = fill

    ws[cell2] = contrat
    ws[cell2].alignment = Alignment(horizontal="center", vertical="center")
    ws[cell2].font = Font(bold=True, size=12,color="000000")
    ws[cell2].fill = fill

def makeEmployeeHoraires(ws,presences,debut):
    fill = PatternFill(start_color=BLEU, end_color=BLEU, fill_type="solid")

    def makeCell(loc, text):
        cell = ws[loc]
        cell.value = text
        cell.alignment = Alignment(horizontal="left", vertical="center")
        cell.fill = fill

    makeCell(f"A{debut}", "Heures normales")
    makeCell(f"A{debut+1}", "Heures supp.")
    makeCell(f"A{debut+2}", "Indem. Trajet Z1")
    makeCell(f"A{debut+3}", "Indem. Trajet Z2")
    makeCell(f"A{debut+4}", "Indem. Trajet Z3")
    makeCell(f"A{debut+5}", "Indem. Trajet Z4")
    makeCell(f"A{debut+6}", "Indem. Trajet Z5")
    makeCell(f"A{debut+7}", "Indem. Transport")
    makeCell(f"A{debut+8}", "Paniers")

    def AddHoraire(col,row, horaire):
        cell = ws[f"{col}{row}"]
        if horaire >= 4:
            cellPanier = ws[f"{col}{row+8}"]
            cellPanier.value = 1
            cellPanier.alignment = Alignment(horizontal="center", vertical="center")

        cell.alignment = Alignment(horizontal="center", vertical="center")
        if cell.value is None:
            cell.value = horaire
        else:
            cell.value += horaire

        if cell.value > 7:
            cellSupp = ws[f"{col}{row+1}"]
            if cellSupp.value is None:
                cellSupp.value = cell.value - 7
            else:
                cellSupp.value += cell.value - 7
            cellSupp.alignment = Alignment(horizontal="center", vertical="center")
            cell.value = 7


    zone_mapping = {"Z1": debut+2, "Z2": debut+3, "Z3": debut+4, "Z4": debut+5, "Z5": debut+6,"z1": debut+2, "z2": debut+3, "z3": debut+4, "z4": debut+5, "z5": debut+6}
    for presence in presences:
        day_num = presence['date'].day
        col_letter = get_column_letter(day_num + 1)
        AddHoraire(col_letter,debut, presence['heures'])
        if presence['zone'] in zone_mapping:
            ws[f"{col_letter}{zone_mapping[presence['zone']]}"].value = 1



    fill = PatternFill(start_color=ORANGE, end_color=ORANGE, fill_type="solid")
    for i in range(debut, debut + 9):
        sum_cell = ws[f"AG{i}"]
        sum_cell.value = f"=ROUND(SUM(B{i}:AF{i}), 2)"
        sum_cell.alignment = Alignment(horizontal="center", vertical="center")
        sum_cell.font = Font(bold=True,color="000000")
        sum_cell.fill = fill
    
    
    fill = PatternFill(start_color=YELLOW, end_color=YELLOW, fill_type="solid")
    ws.merge_cells(f"AH{debut}:AH{debut+2}")
    ws.merge_cells(f"AH{debut+3}:AH{debut+5}")
    ws.merge_cells(f"AH{debut+6}:AH{debut+8}")


    ws[f"AH{debut}"].alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    ws[f"AH{debut}"].font = Font(bold=False,color="000000",size=10)
    ws[f"AH{debut}"].fill = fill

    ws[f"AH{debut+3}"].alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    ws[f"AH{debut+3}"].font = Font(bold=False,color=RED,size=10)
    ws[f"AH{debut+3}"].fill = fill

    ws[f"AH{debut+6}"].alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    ws[f"AH{debut+6}"].font = Font(bold=False,color="000000",size=10)
    ws[f"AH{debut+6}"].fill = fill

   

    return debut + 9

def MakeExcel(infos):
    wb = Workbook()
    ws = wb.active
    ws.title = "Salaries Chantiers"
    makeDate(ws)
    makeTitle(ws,"SALARIES CHANTIER")
    begin = 5
    for key, value in infos.items():
        if 'name' in value and 'lastname' in value and 'contrat' in value:
            name = f"{str(value['lastname']).upper()} {value['name']}"
            contrat = value['contrat']
            makeEmployeeName(ws,name,contrat,begin)
            begin = makeEmployeeHoraires(ws,value['presences'],begin + 2)

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output

if __name__ == "__main__":
    MakeExcel([])